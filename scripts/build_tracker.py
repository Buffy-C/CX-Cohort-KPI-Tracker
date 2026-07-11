"""
KPI Tracker Builder — cohort & month aware
===========================================
Assembles raw CSV exports into a structured Excel tracker.
Targets come from tracker_config.py — the ONLY place targets live.

Usage (run from the project root folder):
    python scripts/build_tracker.py --cohort 1 --wc 2026-06-29
        -> reads  Cohort_1/WC_2026-06-29/raw_data/
        -> writes Cohort_1/WC_2026-06-29/KPI_Tracker_WC_2026-06-29.xlsx

Options:
    --data ./some/folder    override the raw CSV folder
    --out  ./file.xlsx      override the output path

Note: bracketed numbers in filenames, e.g. "agent_productivity (3).csv",
are handled automatically — no need to rename downloads.
"""

import argparse
import re
import sys

# Force UTF-8 output. Windows' default console codepage (cp1252) can't
# encode characters this script prints (≥, ✓, ✗, →, ·), which crashes with
# a UnicodeEncodeError — this happens whether run directly in a plain
# cmd.exe/PowerShell window or via a wrapper script that captures output.
# reconfigure() is available on Python 3.7+; the getattr guard is just
# defensive for any oddball stdout stream that doesn't support it.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tracker_config as cfg

# Populated in main() from tracker_config — do not hard-code values here.
TARGETS = {}
MONTH = None
TAG = ""  # provenance: "Cohort 1 · W/C 22 Jun 2026 · Month 3 Targets"
QA_PASS_RATE_THRESHOLD = cfg.QA_PASS_RATE_THRESHOLD

# ── So Energy brand colours ───────────────────────────────────────────────────
C_LIME    = "CBFB00"
C_BLACK   = "131409"
C_DARK    = "242C28"
C_WHITE   = "FFFFFF"
C_TEAL    = "00D9AF"
C_RED     = "FF3F46"
C_BLUE    = "00BBFF"
C_GREY    = "81877C"
C_LGREY   = "F5F5F5"
C_GRID    = "E8E7E1"
C_AMBER   = "F5C400"

FILL_HEADER  = PatternFill("solid", fgColor=C_BLACK)
FILL_SUBHDR  = PatternFill("solid", fgColor=C_DARK)
FILL_LIME    = PatternFill("solid", fgColor=C_LIME)
FILL_TEAL    = PatternFill("solid", fgColor="CCF7EE")
FILL_AMBER   = PatternFill("solid", fgColor="FFF5B0")
FILL_RED     = PatternFill("solid", fgColor="FFE5E6")
FILL_GREY_BG = PatternFill("solid", fgColor=C_LGREY)
FILL_WHITE   = PatternFill("solid", fgColor=C_WHITE)
FILL_ALT     = PatternFill("solid", fgColor="F9F9F9")
FILL_DARK    = PatternFill("solid", fgColor=C_DARK)

FONT_HDR    = Font(name="Arial", bold=True, color=C_WHITE, size=10)
FONT_SUBHDR = Font(name="Arial", bold=True, color=C_LIME,  size=9)
FONT_LABEL  = Font(name="Arial", bold=True, color=C_BLACK, size=9)
FONT_BODY   = Font(name="Arial", color=C_BLACK, size=9)
FONT_LIME   = Font(name="Arial", bold=True, color=C_BLACK, size=9)

ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN  = Side(style="thin",   color=C_GRID)
BORDER_THIN  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Helpers ───────────────────────────────────────────────────────────────────

def fmt_tgt(key):
    """Format a target for display; None -> 'tracked'."""
    v = TARGETS.get(key)
    if v is None:
        return "tracked (no target)"
    if key == "RPH":
        return f"{v}"
    if key in ("QA Score", "Compliance"):
        return f"{v:.0f}%"
    return f"{v:.0%}"


def parse_qa_agent(s):
    """'Shania.Solomons@so.energy (Solomons, Shania)' -> 'Shania Solomons'
    Returns None for blank/missing values — the QA and Compliance exports
    include a per-evaluation-form subtotal row with an empty Agent field,
    and without this check that row gets stringified to the literal text
    "nan" and treated as a phantom agent."""
    if pd.isna(s) or str(s).strip() == "":
        return None
    s = str(s)
    m = re.search(r'\(([^,]+),\s*([^)]+)\)', s)
    if m:
        return f"{m.group(2).strip()} {m.group(1).strip()}"
    if '@' in s:
        local = s.split('@')[0]
        return ' '.join(p.capitalize() for p in local.split('.'))
    return s

def email_to_name(email):
    local = str(email).split('@')[0]
    return ' '.join(p.capitalize() for p in local.split('.'))

def parse_time(t):
    """'00:14:20' -> total minutes (float)"""
    if pd.isna(t) or str(t).strip() == '':
        return np.nan
    try:
        parts = str(t).strip().split(':')
        h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
        return round(h * 60 + m + s / 60, 2)
    except Exception:
        return np.nan

def pct_to_float(v):
    if pd.isna(v):
        return np.nan
    s = str(v).replace('%', '').strip()
    try:
        return float(s) / 100
    except Exception:
        return np.nan

def strip_pct_sign(v):
    return str(v).lstrip("'")


# ── Load & merge all sources ──────────────────────────────────────────────────

def find_file(data_dir: Path, name: str) -> Path:
    """Find a file tolerating spaces vs underscores and bracketed numbers."""
    d = data_dir
    if (d / name).exists():
        return d / name
    spaced = name.replace("_", " ")
    if (d / spaced).exists():
        return d / spaced
    stem = Path(name).stem.lower().replace("_", " ").replace("-", " ")
    stem_clean = re.sub(r"[\s_]*[\(\[]?\d+[\)\]]?[\s_]*", " ", stem).strip()
    for f in d.iterdir():
        if f.suffix.lower() != Path(name).suffix.lower():
            continue
        candidate = f.stem.lower().replace("_", " ").replace("-", " ")
        candidate_clean = re.sub(r"[\s_]*[\(\[]?\d+[\)\]]?[\s_]*", " ", candidate).strip()
        if candidate_clean == stem_clean:
            return f
    raise FileNotFoundError(f"Could not find '{name}' or equivalent in {d}")


def load_all(data_dir: Path) -> dict:
    results = {}

    def ff(name: str) -> Path:
        return find_file(data_dir, name)

    # 1. Productivity
    prod = pd.read_csv(ff("agent_productivity.csv"), index_col=0)
    prod.columns = prod.columns.str.strip()
    prod = prod.rename(columns={
        "CSAT Results - Blended Customer Satisfaction %": "CSAT_blended",
        "First Contact Resolution %": "FCR_str",
        "Resolves per Hour": "RPH",
        "Number of Contacts Handled": "Contacts",
        "Number of Contacts Disconnected": "Disconnected",
        "Contacts per Hour": "CPH",
        "Number of Closed Cases": "Closed Cases",
        "Number of Open Cases": "Open Cases",
        "Productive Hours": "Prod Hours",
    })
    prod["FCR"] = prod["FCR_str"].apply(pct_to_float)
    prod["QA Score (prod)"] = pd.to_numeric(prod["QA Score"], errors="coerce")
    prod = prod.drop(columns=["QA Score"])
    results["prod"] = prod

    # 2. Performance (handle times)
    perf = pd.read_csv(ff("agent_performance.csv"), index_col=0)
    perf.columns = perf.columns.str.strip()
    perf = perf[perf["Agent Name"].notna()].copy()
    for col, new in [
        ("Average Handle Time - All Channels", "AHT_raw"),
        ("Average ACW Time - All Channels",    "ACW_raw"),
        ("Average Hold Time - All Channels",   "Hold_raw"),
        ("Average Speed of Answer  - Live Channels", "ASA_raw"),
    ]:
        if col in perf.columns:
            perf[new] = perf[col]
    perf["AHT_min"] = perf["AHT_raw"].apply(parse_time)
    perf["ACW_min"] = perf["ACW_raw"].apply(parse_time)
    perf["Hold_min"]= perf["Hold_raw"].apply(parse_time)
    perf["ASA_min"] = perf["ASA_raw"].apply(parse_time)
    perf = perf.rename(columns={
        "Number of Inbound Calls Answered": "Inbound Calls",
        "Number of Email Responses":        "Emails",
        "Number of Outbound Calls":         "Outbound Calls",
        "Number of Outbound Emails":        "Outbound Emails",
        "Number of Callbacks Answered":     "Callbacks",
    })
    results["perf"] = perf

    # 3. QA scores
    qa = pd.read_csv(ff("QA Dashboard-Quality Scores.csv"))
    qa.columns = qa.columns.str.strip()
    qa["Agent Name"] = qa["Agent"].apply(parse_qa_agent)
    qa = qa[qa["Agent Name"].notna()].copy()   # drop per-form subtotal rows (blank Agent)
    qa["QA Score"]   = pd.to_numeric(qa["Avg. evaluation score"], errors="coerce")
    qa["QA Evals"]   = pd.to_numeric(qa["Evaluations performed"], errors="coerce")
    qa["QA Prior"]   = pd.to_numeric(qa["Prior avg. evaluation score"], errors="coerce")
    qa["QA Change"]  = qa["Percent change in avg. evaluation score"].apply(strip_pct_sign)

    def _wavg(g, val_col, wt_col):
        d = g[[val_col, wt_col]].dropna()
        return (d[val_col] * d[wt_col]).sum() / d[wt_col].sum() if d[wt_col].sum() else np.nan

    qa = qa.groupby("Agent Name").apply(lambda g: pd.Series({
        # An agent can appear on more than one row if evaluated on more than
        # one channel (e.g. Call + Email) in the same week. Weight by evals
        # per row rather than taking a plain mean across rows, so an agent
        # with 5 call evals and 1 email eval isn't blended 50/50.
        "QA Score": _wavg(g, "QA Score", "QA Evals"),
        "QA Evals": g["QA Evals"].sum(min_count=1),
        "QA Prior": g["QA Prior"].mean(),
        "QA Change": g["QA Change"].iloc[0],
    }), include_groups=False).reset_index()
    results["qa"] = qa[["Agent Name","QA Score","QA Evals","QA Prior","QA Change"]]

    # 4. Compliance scores
    comp = pd.read_csv(ff("QA Dashboard-Compliance scorecard.csv"))
    comp.columns = comp.columns.str.strip()
    comp["Agent Name"] = comp["Agent"].apply(parse_qa_agent)
    comp = comp[comp["Agent Name"].notna()].copy()   # drop per-form subtotal rows (blank Agent)
    comp["Compliance"] = pd.to_numeric(comp["Avg. evaluation score"], errors="coerce")
    comp["Comp Evals"] = pd.to_numeric(comp["Evaluations performed"], errors="coerce")
    comp["Comp Prior"] = pd.to_numeric(comp["Prior avg. evaluation score"], errors="coerce")
    comp["Comp Change"] = comp["Percent change in avg. evaluation score"].apply(strip_pct_sign)
    comp = comp.groupby("Agent Name").apply(lambda g: pd.Series({
        "Compliance": _wavg(g, "Compliance", "Comp Evals"),
        "Comp Evals": g["Comp Evals"].sum(min_count=1),
        "Comp Prior": g["Comp Prior"].mean(),
        "Comp Change": g["Comp Change"].iloc[0],
    }), include_groups=False).reset_index()
    results["comp"] = comp[["Agent Name","Compliance","Comp Evals","Comp Prior","Comp Change"]]

    # 5. Complaints closed by agent (fallback summary)
    try:
        cc = pd.read_csv(ff("complaints_closed_by_agent.csv"), index_col=0)
        cc.columns = cc.columns.str.strip()
        cc["Agent Name"] = cc["Last Complaint Closure Agent"].apply(email_to_name)
        results["cc"] = cc[["Agent Name","Complaints Closed"]]
    except Exception:
        results["cc"] = pd.DataFrame(columns=["Agent Name","Complaints Closed"])

    # 6. CSAT — blended (average of Agent Rating % Positive and Resolution Rating % Positive)
    try:
        csat_raw = pd.read_csv(ff("Blended_Customer_Satisfaction__.csv"), index_col=0)
        csat_raw.columns = csat_raw.columns.str.strip()
        csat_raw["Positive_Agent"]      = csat_raw["CSAT Agent Rating Category"].str.strip() == "Positive"
        csat_raw["Positive_Resolution"] = csat_raw["CSAT Resolution Rating Category"].str.strip() == "Positive"
        csat_agg = csat_raw.groupby("Agent Name").agg(
            Surveys=("Contact ID","count"),
            Positive_Agent=("Positive_Agent","sum"),
            Positive_Resolution=("Positive_Resolution","sum")).reset_index()
        csat_agg["CSAT_Agent_Rating"]      = csat_agg["Positive_Agent"] / csat_agg["Surveys"]
        csat_agg["CSAT_Resolution_Rating"] = csat_agg["Positive_Resolution"] / csat_agg["Surveys"]
        # Blended CSAT = simple average of the two % Positive streams (matches team-level
        # "blended_customer_satisfaction__.csv" export methodology)
        csat_agg["CSAT"] = (csat_agg["CSAT_Agent_Rating"] + csat_agg["CSAT_Resolution_Rating"]) / 2
        results["csat_agent"] = csat_agg[
            ["Agent Name","CSAT","CSAT_Agent_Rating","CSAT_Resolution_Rating","Surveys"]]

        # Pooled (volume-weighted) team-level blend — sums raw counts across ALL
        # surveys before dividing, rather than averaging the per-agent CSAT %s.
        # This is what reconciles to the platform's own team-level CSAT figure;
        # a plain mean of agent CSATs understates the score whenever agents have
        # uneven survey volumes (small-n agents get the same weight as big-n ones).
        total_surveys = int(csat_raw["Contact ID"].count())
        if total_surveys:
            pct_agent      = csat_raw["Positive_Agent"].sum() / total_surveys
            pct_resolution = csat_raw["Positive_Resolution"].sum() / total_surveys
            results["csat_team"] = {
                "Surveys": total_surveys,
                "CSAT_Agent_Rating": pct_agent,
                "CSAT_Resolution_Rating": pct_resolution,
                "CSAT": (pct_agent + pct_resolution) / 2,
            }
        else:
            results["csat_team"] = None
    except Exception:
        results["csat_agent"] = None
        results["csat_team"] = None

    # 7. D1 / D28 / D56 from closed complaints
    try:
        comp_raw = pd.read_csv(ff("Number of Closed Complaints.csv"), index_col=0)
        comp_raw.columns = comp_raw.columns.str.strip()

        def extract_agent(row):
            closure = row.get("Complaint Closure Agent")
            created = row.get("Complaint Created Agent")
            if pd.notna(closure) and "@" in str(closure):
                return email_to_name(closure)
            if pd.notna(created) and "@" in str(created):
                return email_to_name(created)
            return None

        comp_raw["Agent Name"] = comp_raw.apply(extract_agent, axis=1)
        comp_raw = comp_raw[comp_raw["Agent Name"].notna()].copy()
        comp_raw["Agent Name"] = comp_raw["Agent Name"].str.strip()
        comp_raw["D1_bool"]  = comp_raw["Resolved within 1 day (Yes / No)"].str.upper() == "YES"
        comp_raw["D28_bool"] = comp_raw["Resolved within 28 days (Yes / No)"].str.upper() == "YES"
        if "Resolved within 56 days (Yes / No)" in comp_raw.columns:
            comp_raw["D56_bool"] = comp_raw["Resolved within 56 days (Yes / No)"].str.upper() == "YES"
        else:
            comp_raw["D56_bool"] = comp_raw["D28_bool"]  # D28 implies D56

        d_agg = comp_raw.groupby("Agent Name").agg(
            Complaints_Closed=("Case ID", "count"),
            D1=("D1_bool",  "mean"),
            D28=("D28_bool", "mean"),
            D56=("D56_bool", "mean")).reset_index()
        results["d1_agent"] = d_agg

        # Team-level pooled D1/D28/D56, computed from EVERY closed complaint
        # attributed to the team — not restricted to agents who happen to
        # appear in this week's agent_productivity.csv. An agent can close a
        # complaint in a week they weren't on live queues (leave, ad-hoc
        # complaint work, etc.), and restricting to the productivity roster
        # silently drops those complaints from the team total, understating
        # the pooled rate. This mirrors the CSAT_team pooling above.
        total_closed = int(comp_raw["Case ID"].count())
        if total_closed:
            results["complaints_team"] = {
                "Complaints Closed": total_closed,
                "D1":  comp_raw["D1_bool"].sum()  / total_closed,
                "D28": comp_raw["D28_bool"].sum() / total_closed,
                "D56": comp_raw["D56_bool"].sum() / total_closed,
            }
        else:
            results["complaints_team"] = None
    except Exception as e:
        print("Warning: D1/D28/D56 load failed:", e)
        results["d1_agent"] = None
        results["complaints_team"] = None

    # 8. Team-level stats
    def read_single(fname):
        try:
            df_tmp = pd.read_csv(ff(fname))
            return str(df_tmp.iloc[0, -1]).strip()
        except Exception:
            return None

    def mean_time(series):
        vals = pd.to_numeric(series, errors="coerce").dropna()
        if vals.empty:
            return "—"
        avg = vals.mean()
        return f"{int(avg//60):02d}:{int(avg%60):02d}:{int((avg*60)%60):02d}"

    perf = results["perf"]
    results["team"] = {
        "CSAT_Pooled":           results.get("csat_team"),
        "Complaints_Pooled":     results.get("complaints_team"),
        "Avg Handle Time":      mean_time(perf.get("AHT_min", pd.Series())),
        "Avg ACW Time":         mean_time(perf.get("ACW_min", pd.Series())),
        "Avg Hold Time":        mean_time(perf.get("Hold_min", pd.Series())),
        "Avg Speed of Answer":  mean_time(perf.get("ASA_min", pd.Series())),
        "Avg Days to Resolve":  read_single("average_days_to_resolve_complaint.csv"),
        "Avg Open Complaint Age (Days)": read_single(
            "average_complaint_age_of_open_complaints__calendar_days_.csv"),
        "Complaints Closed (week)": read_single("complaint_closure_volumes.csv"),
    }
    return results


def build_agent_df(results: dict) -> pd.DataFrame:
    prod = results["prod"].copy()
    perf = results["perf"].copy()

    df = prod.merge(
        perf[["Agent Name","Inbound Calls","Emails","Outbound Calls",
              "Outbound Emails","Callbacks","AHT_min","ACW_min","Hold_min","ASA_min"]],
        on="Agent Name", how="left")
    df = df.merge(results["qa"],   on="Agent Name", how="left")
    df = df.merge(results["comp"], on="Agent Name", how="left")
    df = df.merge(results["cc"],   on="Agent Name", how="left")

    if results.get("csat_agent") is not None:
        df = df.merge(results["csat_agent"], on="Agent Name", how="left")
    else:
        df["CSAT"] = np.nan
        df["Surveys"] = np.nan

    if results.get("d1_agent") is not None:
        df = df.merge(
            results["d1_agent"][["Agent Name","D1","D28","D56","Complaints_Closed"]],
            on="Agent Name", how="left")
        df["Complaints Closed"] = df["Complaints_Closed"].fillna(df["Complaints Closed"])
        df = df.drop(columns=["Complaints_Closed"])
    else:
        df["D1"] = df["D28"] = df["D56"] = np.nan

    # Filter out Team Leaders
    tl_names = set(df["Team Leader"].dropna().unique())
    df = df[~df["Agent Name"].isin(tl_names)].copy().reset_index(drop=True)
    return df


# ── Excel writer helpers ──────────────────────────────────────────────────────

def style_cell(cell, fill=None, font=None, align=None, border=None, num_fmt=None):
    if fill:    cell.fill      = fill
    if font:    cell.font      = font
    if align:   cell.alignment = align
    if border:  cell.border    = border
    if num_fmt: cell.number_format = num_fmt

def write_section_header(ws, row, col_start, col_end, label, fill=None, font=None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=label)
    style_cell(cell, fill=fill or FILL_SUBHDR, font=font or FONT_SUBHDR,
               align=ALIGN_C, border=BORDER_THIN)
    ws.row_dimensions[row].height = 20


# ── Sheet 0: Control / File Index ────────────────────────────────────────────

def write_control_sheet(wb, data_dir: Path):
    ws = wb.create_sheet("Control", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    style_cell(ws.cell(row=1, column=1,
               value=f"KPI Tracker — Data Sources & Control  |  {TAG}"),
               fill=FILL_HEADER,
               font=Font(name="Arial", bold=True, color=C_LIME, size=14),
               align=ALIGN_C)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    style_cell(ws.cell(row=2, column=1,
               value="Required: agent_productivity.csv  |  agent_performance.csv  |  Number of Closed Complaints.csv  |  Blended_Customer_Satisfaction__.csv  |  QA Dashboard-Quality Scores.csv  |  QA Dashboard-Compliance scorecard.csv     Run: python scripts/build_tracker.py --cohort N --wc YYYY-MM-DD"),
               fill=FILL_GREY_BG,
               font=Font(name="Arial", italic=True, color=C_GREY, size=9),
               align=ALIGN_L)
    ws.row_dimensions[2].height = 18

    widths = [30, 32, 46, 12, 12, 20, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ["Filename", "Source / Report Name", "How It's Calculated",
                "Required?", "Present?", "Populates", "Notes"]
    ws.row_dimensions[3].height = 24
    for c, h in enumerate(headers, 1):
        style_cell(ws.cell(row=3, column=c, value=h),
                   fill=FILL_SUBHDR, font=FONT_SUBHDR, align=ALIGN_C, border=BORDER_THIN)

    FILES = [
        ("agent_productivity.csv",
         "Agent Productivity report",
         "RPH, FCR %, contact/case counts and Productive Hours are taken directly "
         "from the platform export as-is — no calculation is applied, other than "
         "parsing the FCR % text into a decimal.",
         True,
         "Agent Tracker — core",
         "RPH, FCR, contacts, hours, open/closed cases"),
        ("agent_performance.csv",
         "Agent Performance report",
         "AHT, ACW, Hold and ASA are exported as HH:MM:SS text and converted to "
         "decimal minutes (hours×60 + minutes + seconds÷60). Call/email volumes "
         "are taken directly from the export.",
         True,
         "Agent Tracker — handle times",
         "AHT, ACW, hold time, ASA, call/email volumes"),
        ("QA Dashboard-Quality Scores.csv",
         "QA Dashboard → Quality Scores",
         "QA Score = mean of 'Avg. evaluation score' across all rows for that "
         "agent (averaged if multiple evaluations/periods are present). "
         "QA Evals = sum of 'Evaluations performed'.",
         True,
         "QA & Compliance tab + Agent Tracker",
         f"Accumulating scores — target {fmt_tgt('QA Score')} "
         f"({TARGETS['QA Autofails']} autofail max)"),
        ("QA Dashboard-Compliance scorecard.csv",
         "QA Dashboard → Compliance Scorecard",
         "Compliance = mean of 'Avg. evaluation score' across all rows for that "
         "agent. Comp Evals = sum of 'Evaluations performed'. Same logic as "
         "QA Score, applied to the compliance scorecard export.",
         True,
         "QA & Compliance tab + Agent Tracker",
         f"Accumulating scores — target {fmt_tgt('Compliance')}"),
        ("Number of Closed Complaints.csv",
         "AWS Complaints → Closed Complaints (detail)",
         "D1/D28/D56 = % of that agent's closed complaints marked 'Yes' for "
         "Resolved within 1/28/56 days (mean of the Yes/No flag per agent). "
         "Complaints Closed = count of Case IDs per agent.",
         True,
         "Agent Tracker — D1 %, D28 %, D56 %",
         f"D1 target {fmt_tgt('D1')} | D28 {fmt_tgt('D28')} | D56 {fmt_tgt('D56')}"),
        ("Blended_Customer_Satisfaction__.csv",
         "CSAT → Blended Customer Satisfaction (detail)",
         "Blended CSAT = average of two % Positive rates per agent: "
         "(Positive Agent Ratings ÷ Surveys) and (Positive Resolution Ratings ÷ "
         "Surveys), then (%Agent + %Resolution) ÷ 2. Surveys = count of Contact IDs.",
         True,
         "Agent Tracker — CSAT %, Surveys",
         f"Per-agent CSAT % and survey count — target {fmt_tgt('CSAT')}. "
         f"Download may append a number, e.g. '(2)', if re-exported — this is "
         f"matched automatically."),
    ]

    for ri, (fname, source, calc, required, populates, notes) in enumerate(FILES):
        r = 4 + ri
        ws.row_dimensions[r].height = 20
        try:
            find_file(data_dir, fname)
            present = True
        except FileNotFoundError:
            present = False
        fill = FILL_WHITE if ri % 2 == 0 else FILL_ALT

        req_label  = "Required"  if required  else "Optional"
        req_fill   = FILL_RED    if required   else FILL_GREY_BG
        pres_label = "✓ Found"   if present    else "✗ Missing"
        pres_fill  = FILL_TEAL   if present    else (FILL_RED if required else FILL_AMBER)

        row_vals = [fname, source, calc, req_label, pres_label, populates, notes]
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 4:
                style_cell(cell, fill=req_fill,
                           font=Font(name="Arial", bold=True, color=C_BLACK, size=9),
                           align=ALIGN_C, border=BORDER_THIN)
            elif c == 5:
                style_cell(cell, fill=pres_fill,
                           font=Font(name="Arial", bold=True, color=C_BLACK, size=9),
                           align=ALIGN_C, border=BORDER_THIN)
            else:
                style_cell(cell, fill=fill,
                           font=FONT_LABEL if c == 1 else FONT_BODY,
                           align=ALIGN_L, border=BORDER_THIN)

    tgt_row = 4 + len(FILES) + 1
    ws.merge_cells(f"A{tgt_row}:G{tgt_row}")
    style_cell(ws.cell(row=tgt_row, column=1, value=cfg.targets_summary(MONTH)),
               fill=FILL_LIME,
               font=Font(name="Arial", bold=True, color=C_BLACK, size=9),
               align=ALIGN_L)
    ws.row_dimensions[tgt_row].height = 18

    leg_row = tgt_row + 1
    ws.merge_cells(f"A{leg_row}:G{leg_row}")
    style_cell(ws.cell(row=leg_row, column=1,
               value="Legend:   ✓ Found = file detected in your data folder at time of build   |   ✗ Missing = file not found (required files will cause errors; optional files are skipped gracefully)"),
               fill=FILL_GREY_BG,
               font=Font(name="Arial", italic=True, color=C_GREY, size=8),
               align=ALIGN_L)
    ws.row_dimensions[leg_row].height = 16

    ws.sheet_properties.tabColor = C_LIME
    return ws


# ── Sheet 1: Agent Tracker ────────────────────────────────────────────────────

def write_agent_sheet(wb, df):
    ws = wb.create_sheet("Agent Tracker")
    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False

    COLS = [
        ("Agent Name",    "Agent",          20, "@"),
        ("Team Leader",   "Team Leader",    18, "@"),
        ("RPH",           "RPH",             8, "0.00"),
        ("FCR",           "FCR %",           8, "0.0%"),
        ("QA Score",      "QA Score",        9, "0.0"),
        ("QA Evals",      "QA\nEvals",       7, "0"),
        ("Compliance",    "Compliance %",   11, "0.0"),
        ("Comp Evals",    "Comp\nEvals",     7, "0"),
        ("Prod Hours",    "Prod\nHours",     8, "0.00"),
        ("Contacts",      "Contacts",        9, "0"),
        ("Disconnected",  "Disconn.",        8, "0"),
        ("CPH",           "CPH",             7, "0.00"),
        ("Closed Cases",  "Closed",          7, "0"),
        ("Open Cases",    "Open",            7, "0"),
        ("Inbound Calls", "Inbound\nCalls",  8, "0"),
        ("Emails",        "Emails",          7, "0"),
        ("Outbound Calls","Out\nCalls",      7, "0"),
        ("Callbacks",     "Callbacks",       8, "0"),
        ("AHT_min",       "AHT\n(min)",      8, "0.0"),
        ("ACW_min",       "ACW\n(min)",      8, "0.0"),
        ("Hold_min",      "Hold\n(min)",     8, "0.0"),
        ("ASA_min",       "ASA\n(min)",      8, "0.0"),
        ("Complaints Closed", "Complaints\nClosed", 10, "0"),
        ("CSAT",          "CSAT %",          8, "0.0%"),
        ("Surveys",       "CSAT\nSurveys",   8, "0"),
        ("D1",            "D1 %",            8, "0.0%"),
        ("D28",           "D28 %",           8, "0.0%"),
        ("D56",           "D56 %",           9, "0.0%"),
    ]

    SECTIONS = [
        (1,  2,  "IDENTITY",          FILL_HEADER,  FONT_HDR),
        (3,  8,  "CORE KPIs",     FILL_LIME,    FONT_LIME),
        (9,  18, "VOLUME & ACTIVITY", FILL_SUBHDR,  FONT_SUBHDR),
        (19, 22, "HANDLE TIMES",      FILL_SUBHDR,  FONT_SUBHDR),
        (23, 23, "COMPLAINTS",        FILL_SUBHDR,  FONT_SUBHDR),
        (24, 28, "CSAT & RESOLUTION", FILL_SUBHDR,  FONT_SUBHDR),
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    title = ws.cell(row=1, column=1,
                    value=f"Agent Performance Tracker  |  {TAG}")
    style_cell(title, fill=FILL_HEADER,
               font=Font(name="Arial", bold=True, color=C_LIME, size=13), align=ALIGN_C)
    ws.row_dimensions[1].height = 32

    for c_start, c_end, label, fill, font in SECTIONS:
        write_section_header(ws, 2, c_start, c_end, label, fill, font)

    ws.row_dimensions[3].height = 36
    for c, (key, label, width, fmt) in enumerate(COLS, 1):
        cell = ws.cell(row=3, column=c, value=label)
        style_cell(cell, fill=FILL_DARK if c <= 2 else FILL_SUBHDR,
                   font=FONT_HDR if c <= 2 else Font(name="Arial", bold=True, color=C_LIME, size=8),
                   align=ALIGN_C, border=BORDER_THIN)
        ws.column_dimensions[get_column_letter(c)].width = width

    # Targets row
    TARGET_KEYS = ["RPH","FCR","QA Score","Compliance","CSAT","D1","D28","D56"]
    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=1, value="▶ Targets").font = Font(name="Arial", bold=True, color=C_GREY, size=8)
    ws.cell(row=4, column=1).alignment = ALIGN_C
    for c, (key, label, width, fmt) in enumerate(COLS, 1):
        cell = ws.cell(row=4, column=c)
        if key in TARGET_KEYS:
            tv = TARGETS.get(key)
            if tv is not None:
                cell.value = tv
                cell.number_format = fmt
            else:
                cell.value = "N/A"
        style_cell(cell, fill=FILL_GREY_BG,
                   font=Font(name="Arial", italic=True, color=C_GREY, size=8),
                   align=ALIGN_C, border=BORDER_THIN)

    # Data rows
    DATA_START = 5
    for ri, (_, row) in enumerate(df.iterrows()):
        excel_row = DATA_START + ri
        fill = FILL_WHITE if ri % 2 == 0 else FILL_ALT
        ws.row_dimensions[excel_row].height = 18
        for c, (key, label, width, fmt) in enumerate(COLS, 1):
            val = row.get(key, None)
            cell = ws.cell(row=excel_row, column=c)
            if pd.isna(val) if not isinstance(val, str) else val == '':
                cell.value = None
            else:
                cell.value = val
            cell.number_format = fmt
            cell.fill = fill
            cell.font = FONT_BODY if c > 2 else FONT_LABEL
            cell.alignment = ALIGN_C if c > 2 else ALIGN_L
            cell.border = BORDER_THIN

    # Conditional formatting driven by config targets
    data_range_end = DATA_START + len(df) - 1
    col_letters = {key: get_column_letter(i+1) for i, (key, *_ ) in enumerate(COLS)}
    for key in TARGET_KEYS:
        target = TARGETS.get(key)
        if target is None:
            continue
        letter = col_letters[key]
        rng = f"{letter}{DATA_START}:{letter}{data_range_end}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThanOrEqual", formula=[str(target)], fill=FILL_TEAL))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="lessThan", formula=[str(target)], fill=FILL_RED))

    ws.sheet_properties.tabColor = C_LIME
    return ws


# ── Sheet 2: Team Overview ────────────────────────────────────────────────────

def write_team_sheet(wb, df, team_stats):
    ws = wb.create_sheet("Team Overview")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value=f"Team Overview  |  {TAG}")
    style_cell(t, fill=FILL_HEADER,
               font=Font(name="Arial", bold=True, color=C_LIME, size=14), align=ALIGN_C)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    style_cell(ws.cell(row=2, column=1, value="CORE KPIs — Cohort Summary"),
               fill=FILL_SUBHDR, font=FONT_SUBHDR, align=ALIGN_C)
    ws.row_dimensions[2].height = 20

    kpi_headers = ["Metric", "Target", "Cohort Avg", "Assessed", "On Target", "Status", "Δ vs Target", "Notes"]
    ws.row_dimensions[3].height = 22
    for c, h in enumerate(kpi_headers, 1):
        style_cell(ws.cell(row=3, column=c, value=h),
                   fill=FILL_DARK, font=FONT_SUBHDR, align=ALIGN_C, border=BORDER_THIN)

    for col, w in zip("ABCDEFGH", [22, 12, 12, 12, 12, 14, 12, 36]):
        ws.column_dimensions[col].width = w

    tp_df = df.copy()
    N = len(tp_df)

    kpi_rows = [
        ("RPH",        "Resolves per Hour",   TARGETS["RPH"],
         lambda v: f"{v:.2f}", f"Target: {fmt_tgt('RPH')}"),
        ("FCR",        "First Contact Res.",  TARGETS["FCR"],
         lambda v: f"{v:.1%}", f"Target: {fmt_tgt('FCR')}"),
        ("QA Score",   "QA Score",            TARGETS["QA Score"],
         lambda v: f"{v:.1f}%",
         f"Target {fmt_tgt('QA Score')} | {TARGETS['QA Autofails']} autofail max | "
         f"Pass rate ≥ {QA_PASS_RATE_THRESHOLD:.0f}% | "
         f"{int(tp_df['QA Evals'].notna().sum())} agents · {int(tp_df['QA Evals'].sum())} evals"),
        ("Compliance", "Compliance",          TARGETS["Compliance"],
         lambda v: f"{v:.1f}%",
         f"Target {fmt_tgt('Compliance')} | "
         f"{int(tp_df['Comp Evals'].notna().sum())} agents · {int(tp_df['Comp Evals'].sum())} evals"),
        ("CSAT",       "CSAT",                TARGETS["CSAT"],
         lambda v: f"{v:.1%}", f"Target: {fmt_tgt('CSAT')}"),
        ("D1",         "D1 Complaints",       TARGETS["D1"],
         lambda v: f"{v:.1%}", f"Target: {fmt_tgt('D1')}  (resolved within 1 day)"),
        ("D28",        "D28 Complaints",      TARGETS["D28"],
         lambda v: f"{v:.1%}", f"Target: {fmt_tgt('D28')}  (resolved within 28 days)"),
        ("D56",        "D56 Complaints",      TARGETS["D56"],
         lambda v: f"{v:.1%}" if not np.isnan(v) else "—",
         f"Target: {fmt_tgt('D56')}  (resolved within 56 days)"),
    ]

    # Exact-reconstruction weights: col -> volume column. Because each of
    # these per-agent %s is itself Numerator ÷ Volume, weighting the mean by
    # Volume reproduces the true pooled team rate exactly (sum(Numerator) ÷
    # sum(Volume)) rather than a plain per-agent average, which silently
    # gives a 1-response agent the same pull as a high-volume agent.
    EXACT_WEIGHT_COL = {
        "RPH": "Prod Hours",
        "D1":  "Complaints Closed",
        "D28": "Complaints Closed",
        "D56": "Complaints Closed",
        "QA Score":   "QA Evals",
        "Compliance": "Comp Evals",
    }
    # FCR has no raw success/fail counts available per agent — only the
    # platform's own pre-computed % — so this is a best-effort approximation
    # (weighted by Closed Cases) rather than an exact pooled reconstruction.
    APPROX_WEIGHT_COL = {"FCR": "Closed Cases"}

    def _pooled_avg(frame, col, wcol):
        d = frame[[col, wcol]].dropna()
        return (d[col] * d[wcol]).sum() / d[wcol].sum() if len(d) and d[wcol].sum() else np.nan

    for r_offset, (col, label, target, fmt_fn, note) in enumerate(kpi_rows, 4):
        ws.row_dimensions[r_offset].height = 20
        series = tp_df[col].dropna() if col in tp_df.columns else pd.Series([], dtype=float)
        avg = series.mean() if len(series) else np.nan

        if col in EXACT_WEIGHT_COL and EXACT_WEIGHT_COL[col] in tp_df.columns:
            pooled_avg = _pooled_avg(tp_df, col, EXACT_WEIGHT_COL[col])
            if not np.isnan(pooled_avg):
                avg = pooled_avg
                note = f"{note}  |  Pooled by {EXACT_WEIGHT_COL[col]} (not a plain agent average)"
        elif col in APPROX_WEIGHT_COL and APPROX_WEIGHT_COL[col] in tp_df.columns:
            pooled_avg = _pooled_avg(tp_df, col, APPROX_WEIGHT_COL[col])
            if not np.isnan(pooled_avg):
                avg = pooled_avg
                note = (f"{note}  |  Weighted by {APPROX_WEIGHT_COL[col]} — approximation, "
                        f"exact per-agent FCR denominator isn't exported")

        if col == "CSAT":
            # Cohort Avg for CSAT is pooled across all surveys (sum of Positive
            # counts ÷ sum of Surveys) so it reconciles to the platform's own
            # team-level blended CSAT figure. A plain mean of agent CSAT %s
            # under-weights agents with more survey responses and can
            # materially understate the true team score.
            pooled = team_stats.get("CSAT_Pooled")
            if pooled:
                avg = pooled["CSAT"]
                note = (f"{note}  |  Pooled across {pooled['Surveys']} surveys "
                        f"(Agent Rating {pooled['CSAT_Agent_Rating']:.1%}, "
                        f"Resolution Rating {pooled['CSAT_Resolution_Rating']:.1%})")

        if col in ("D1", "D28", "D56"):
            # Prefer the full-team pool (every complaint closed under this
            # team this week) over the roster-restricted pool above. An
            # agent can close a complaint in a week they weren't on live
            # queues at all (leave, ad-hoc complaint handling), so
            # restricting to this week's productivity roster silently drops
            # some of the team's own complaints from the total. This gets
            # closer to — but is not guaranteed to exactly match — the
            # platform's own figure; a residual gap can remain if the
            # platform's D1/D28/D56 widgets apply a filter (e.g. an
            # "Is Autologged" toggle) that isn't present in this export.
            pooled_c = team_stats.get("Complaints_Pooled")
            if pooled_c and col in pooled_c:
                avg = pooled_c[col]
                note = (f"{note}  |  Pooled across all {pooled_c['Complaints Closed']} "
                        f"complaints closed under this team (not restricted to this "
                        f"week's productivity roster) — may still differ slightly from "
                        f"the platform if it applies filters not present in this export")

        if target is None:
            status = "TRACKED"
            status_fill = FILL_GREY_BG
            delta_str = on_tgt_str = "—"
        else:
            n_valid    = len(series)
            n_meet     = int((series >= target).sum()) if n_valid else 0
            delta      = avg - target if not np.isnan(avg) else np.nan
            on_tgt_str = f"{n_meet}/{n_valid}" if n_valid else "—"
            delta_str  = (f"+{fmt_fn(delta)}" if not np.isnan(delta) and delta >= 0
                          else fmt_fn(delta) if not np.isnan(delta) else "—")
            status = ("ON TRACK"     if not np.isnan(avg) and avg >= target else
                      "AT RISK"      if not np.isnan(avg) and avg >= target * 0.85 else
                      "BELOW TARGET" if not np.isnan(avg) else "NO DATA")
            status_fill = (FILL_TEAL  if status == "ON TRACK"     else
                           FILL_AMBER if status == "AT RISK"       else
                           FILL_RED   if status == "BELOW TARGET"  else FILL_GREY_BG)

        target_str = fmt_fn(target) if target is not None else "N/A"
        avg_str    = fmt_fn(avg) if not np.isnan(avg) else "—"
        assessed   = f"{len(series)}/{N}"

        row_data = [label, target_str, avg_str, assessed,
                    on_tgt_str if target is not None else "—",
                    status, delta_str, note]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_offset, column=c, value=val)
            is_status = (c == 6)
            style_cell(cell,
                       fill=status_fill if is_status else (FILL_WHITE if r_offset % 2 else FILL_ALT),
                       font=Font(name="Arial", bold=is_status, color=C_BLACK, size=9),
                       align=ALIGN_C, border=BORDER_THIN)

    # Operational metrics
    op_row = 4 + len(kpi_rows) + 1
    ws.merge_cells(f"A{op_row}:H{op_row}")
    style_cell(ws.cell(row=op_row, column=1, value="OPERATIONAL METRICS — Team Level"),
               fill=FILL_SUBHDR, font=FONT_SUBHDR, align=ALIGN_C)
    ws.row_dimensions[op_row].height = 20

    op_headers = ["Metric", "Value", "Source", "", "Metric", "Value", "Source", ""]
    ws.row_dimensions[op_row+1].height = 20
    for c, h in enumerate(op_headers, 1):
        style_cell(ws.cell(row=op_row+1, column=c, value=h),
                   fill=FILL_DARK, font=FONT_SUBHDR, align=ALIGN_C, border=BORDER_THIN)

    op_items = [
        ("Avg Handle Time",          team_stats.get("Avg Handle Time","—"),       "agent_performance.csv"),
        ("Avg ACW Time",             team_stats.get("Avg ACW Time","—"),          "agent_performance.csv"),
        ("Avg Hold Time",            team_stats.get("Avg Hold Time","—"),         "agent_performance.csv"),
        ("Avg Speed of Answer",      team_stats.get("Avg Speed of Answer","—"),   "agent_performance.csv"),
        ("Avg Days to Resolve",      team_stats.get("Avg Days to Resolve","—"),   "average_days_to_resolve_complaint.csv"),
        ("Avg Open Complaint Age",   team_stats.get("Avg Open Complaint Age (Days)","—"), "average_complaint_age__calendar_days_.csv"),
        ("Complaints Closed (week)", team_stats.get("Complaints Closed (week)","—"), "complaint_closure_volumes.csv"),
    ]

    for i, (metric, val, src) in enumerate(op_items):
        r = op_row + 2 + (i // 2)
        c_offset = 0 if i % 2 == 0 else 4
        ws.row_dimensions[r].height = 18
        fill = FILL_WHITE if (r % 2 == 0) else FILL_ALT
        for c, v in enumerate([metric, val, src, ""], 1):
            cell = ws.cell(row=r, column=c + c_offset, value=v)
            style_cell(cell, fill=fill,
                       font=FONT_LABEL if c == 1 else FONT_BODY,
                       align=ALIGN_L if c in (1, 3) else ALIGN_C,
                       border=BORDER_THIN)

    # Quadrant summary
    q_row = op_row + 2 + (len(op_items) + 1) // 2 + 2

    def classify(row):
        fast    = pd.notna(row["RPH"]) and row["RPH"] >= TARGETS["RPH"]
        quality = pd.notna(row["FCR"]) and row["FCR"] >= TARGETS["FCR"]
        if fast and quality:      return "High Performance"
        if not fast and quality:  return "Experience Focus"
        if fast and not quality:  return "Speed Risk"
        return "Underperforming"

    tp_df["Quadrant"] = tp_df.apply(classify, axis=1)
    q_counts = tp_df["Quadrant"].value_counts()

    ws.merge_cells(f"A{q_row}:H{q_row}")
    style_cell(ws.cell(row=q_row, column=1, value="PERFORMANCE QUADRANT SUMMARY"),
               fill=FILL_SUBHDR, font=FONT_SUBHDR, align=ALIGN_C)
    ws.row_dimensions[q_row].height = 20

    q_fills = {
        "High Performance": FILL_TEAL,
        "Experience Focus": PatternFill("solid", fgColor="E0F8FF"),
        "Speed Risk":       FILL_AMBER,
        "Underperforming":  FILL_RED,
    }
    for i, q in enumerate(["High Performance", "Experience Focus", "Speed Risk", "Underperforming"]):
        r = q_row + 1 + i
        ws.row_dimensions[r].height = 18
        agents = tp_df[tp_df["Quadrant"] == q]["Agent Name"].tolist()
        for c, val in enumerate([q, str(q_counts.get(q, 0)), ", ".join(agents), ""], 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_cell(cell, fill=q_fills[q],
                       font=Font(name="Arial", bold=(c == 1), color=C_BLACK, size=9),
                       align=ALIGN_L if c in (1, 3) else ALIGN_C, border=BORDER_THIN)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)

    ws.sheet_properties.tabColor = C_TEAL
    return ws


# ── Sheet 3: QA & Compliance detail ──────────────────────────────────────────

def write_qa_sheet(wb, results):
    ws = wb.create_sheet("QA & Compliance")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    style_cell(ws.cell(row=1, column=1,
               value=f"QA & Compliance — Agent Detail  |  {TAG}  |  QA target {fmt_tgt('QA Score')}  |  Compliance target {fmt_tgt('Compliance')}"),
               fill=FILL_HEADER, font=Font(name="Arial", bold=True, color=C_LIME, size=13),
               align=ALIGN_C)
    ws.row_dimensions[1].height = 32

    qa   = results["qa"].copy()
    comp = results["comp"].copy()
    merged = qa.merge(comp, on="Agent Name", how="outer", suffixes=("_qa","_comp"))

    headers = [
        ("Agent Name",   "Agent",              22, "@"),
        ("QA Score",     "QA Score %",          12, "0.00"),
        ("QA Evals",     "QA Evals",             9, "0"),
        ("QA Prior",     "Prior QA %",          12, "0.00"),
        ("QA Change",    "QA Δ %",              10, "@"),
        ("Compliance",   "Compliance %",        14, "0.00"),
        ("Comp Evals",   "Comp Evals",           9, "0"),
        ("Comp Prior",   "Prior Comp %",        14, "0.00"),
        ("Comp Change",  "Comp Δ %",            10, "@"),
    ]

    ws.merge_cells("A2:E2")
    style_cell(ws.cell(row=2, column=1, value=f"QUALITY SCORES  (target ≥ {fmt_tgt('QA Score')})"),
               fill=FILL_LIME, font=FONT_LIME, align=ALIGN_C)
    ws.merge_cells("F2:I2")
    style_cell(ws.cell(row=2, column=6, value=f"COMPLIANCE SCORECARD  (target ≥ {fmt_tgt('Compliance')})"),
               fill=FILL_SUBHDR, font=FONT_SUBHDR, align=ALIGN_C)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 22
    for c, (key, label, width, fmt) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=label)
        style_cell(cell, fill=FILL_DARK, font=FONT_SUBHDR, align=ALIGN_C, border=BORDER_THIN)
        ws.column_dimensions[get_column_letter(c)].width = width

    for ri, (_, row) in enumerate(merged.iterrows()):
        r = 4 + ri
        ws.row_dimensions[r].height = 18
        fill = FILL_WHITE if ri % 2 == 0 else FILL_ALT
        for c, (key, *_) in enumerate(headers, 1):
            val = row.get(key, None)
            cell = ws.cell(row=r, column=c, value=val if not pd.isna(val) else None)
            cell.number_format = headers[c-1][3]
            style_cell(cell, fill=fill,
                       font=FONT_LABEL if c == 1 else FONT_BODY,
                       align=ALIGN_L if c == 1 else ALIGN_C, border=BORDER_THIN)

    n_data = len(merged)
    for col_letter, key in [("B", "QA Score"), ("F", "Compliance")]:
        ws.conditional_formatting.add(f"{col_letter}4:{col_letter}{3+n_data}", CellIsRule(
            operator="greaterThanOrEqual", formula=[str(TARGETS[key])], fill=FILL_TEAL))
        ws.conditional_formatting.add(f"{col_letter}4:{col_letter}{3+n_data}", CellIsRule(
            operator="lessThan", formula=[str(TARGETS[key])], fill=FILL_RED))

    ws.sheet_properties.tabColor = C_GREY
    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    global TARGETS, MONTH, TAG

    parser = argparse.ArgumentParser(description="Build KPI tracker from CSV exports")
    parser.add_argument("--cohort", type=int, required=True,
                        help=f"Cohort number: {sorted(cfg.COHORTS)}")
    parser.add_argument("--wc", default=None, help="Weekly mode: W/C date")
    parser.add_argument("--month", type=int, default=None, help="Full-month mode")
    parser.add_argument("--data", default=None,
                        help="Override raw CSV folder (default: Cohort_N/WC_date/raw_data)")
    parser.add_argument("--out", default=None, help="Override output filename")
    args = parser.parse_args()

    if bool(args.wc) == bool(args.month):
        sys.exit("Pass exactly one of --wc or --month.")
    cohort_name = cfg.COHORTS[args.cohort]["name"]
    root = Path(__file__).resolve().parent.parent
    if args.month:
        MONTH = args.month
        TARGETS = cfg.TARGETS_BY_MONTH[MONTH]
        week_dir = root / f"Cohort_{args.cohort}" / f"Month_{MONTH}"
        stamp = f"C{args.cohort}_Month_{MONTH}"
        TAG = f"{cohort_name} · Month {MONTH} (full month) · Month {MONTH} Targets"
        period_desc = f"Month {MONTH} (full month)"
    else:
        MONTH, TARGETS = cfg.get_targets(args.cohort, args.wc)
        week_dir = root / f"Cohort_{args.cohort}" / f"WC_{args.wc}"
        stamp = f"C{args.cohort}_WC_{args.wc}"
        TAG = f"{cfg.week_label(args.cohort, args.wc)} · Month {MONTH} Targets"
        period_desc = f"W/C {args.wc}"
    data_dir = Path(args.data) if args.data else week_dir / "raw_data"
    if not data_dir.exists():
        sys.exit(f"Raw data folder not found: {data_dir}\n"
                 f"Create it and drop this week's CSV exports inside, then re-run.")

    out = Path(args.out) if args.out else week_dir / f"KPI_Tracker_{stamp}.xlsx"

    print(f"Cohort:  {cohort_name} (started {cfg.COHORTS[args.cohort]['start']})")
    print(f"Period:  {period_desc}  ->  Month {MONTH} targets")
    print(f"Targets: {cfg.targets_summary(MONTH)}")
    print(f"Data:    {data_dir.resolve()}")

    results = load_all(data_dir)
    df      = build_agent_df(results)
    print(f"Loaded {len(df)} TP agents")

    wb = Workbook()
    wb.remove(wb.active)
    write_control_sheet(wb, data_dir)
    write_team_sheet(wb, df, results["team"])
    write_agent_sheet(wb, df)
    write_qa_sheet(wb, results)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Saved -> {out.resolve()}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
